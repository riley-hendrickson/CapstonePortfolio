import itertools
import tkinter as tk
from tkinter import ttk
from dataclasses import dataclass, field
from spreadsheets import Spreadsheet
from openpyxl.utils.cell import get_column_letter
from pathlib import Path
import re

# TODO: display sample row values using the selected indices?
# TODO: replace color leasing with fixed colors in order for the indices?

SELECTION_COLORS = [
    '#0066ff', # blue
    '#009966', # green
    '#ff9900', # orange
    '#cc0000', # red
    '#9933cc', # purple
]

@dataclass
class IndexInfo:
    key: str
    description: str
    defaultIndex: int = -1
    patterns: list[str] = field(default_factory=list) # regex patterns for auto-detection

INDICES_INFO = [
    IndexInfo('firstName', 'First Name', patterns=[r'first.?name']),
    IndexInfo('lastName', 'Last Name', patterns=[r'last.?name']),
    IndexInfo('gradYear', 'Graduation Year', patterns=[r'(?:grad(?:uation)?|degree).?year', r'class', r'year']),
    IndexInfo('email', 'Email (to filter out empty rows)', patterns=[r'e[- ]?mail']),
]

class ColumnSelectGUI:
    def __init__(self, inputPath, indices: list[IndexInfo] = INDICES_INFO):
        self.result = None

        self.window = tk.Toplevel()
        self.window.title("Confirm column indices")
        self.window.resizable(False, False)
        self.window.attributes('-topmost', True)
        self.window.grab_set() # prevent interaction with other windows

        if isinstance(inputPath, str):
            inputPath = Path(inputPath)
        elif not isinstance(inputPath, Path):
            raise TypeError('inputPath must be a string or Path')

        self.titleLabel = tk.Label(self.window, text='Select input columns:')
        self.titleLabel.config(font='-weight bold -size 14')
        self.titleLabel.pack(pady=(10, 0))
        self.filePathLabel = tk.Label(self.window, text=inputPath)
        self.filePathLabel.pack(pady=(0, 10))

        inputSpreadsheet = Spreadsheet.fromFile(inputPath)
        self.preview = PreviewFrame(inputSpreadsheet, self)
        self.preview.pack(padx=10, pady=10, fill=tk.X, expand=True)

        self.confirmBtn = ttk.Button(self.window, text='Confirm selections', command=self.onConfirmBtnClick)

        self.selectTarget = None

        self.selectionFrames = []
        for index in indices:
            i = index.defaultIndex
            if i < 0:
                def findIndex(regex):
                    return next((i for i, h in enumerate(inputSpreadsheet.getHeaders()) if re.match(regex, str(h), re.IGNORECASE)), -1)
                for pattern in index.patterns:
                    i = findIndex(pattern)
                    if i >= 0:
                        break
            if i >= 0:
                print('Auto-selected column', i, 'for', index.key)
                if self.preview.isSelected(i):
                    i = -1
                else:
                    self.preview.select(i)
            frame = IndexSelectionFrame(index, i, self)
            frame.pack(pady=2)
            self.selectionFrames.append(frame)

        self.checkButtonEnabled()
        self.confirmBtn.pack(pady=10)

        #self.window.mainloop()

    def checkButtonEnabled(self):
        enabled = all(frame.columnIndex >= 0 for frame in self.selectionFrames)
        self.confirmBtn.state(['!disabled'] if enabled else ['disabled'])

    def onConfirmBtnClick(self):
        self.confirmBtn.state(['disabled'])
        indices = { frame.indexInfo.key: frame.columnIndex for frame in self.selectionFrames }
        print('CONFIRMED INDICES:', indices)
        self.result = indices
        self.window.destroy()

    def startSelect(self, target):
        if self.selectTarget is not None:
            self.selectTarget.cancelSelect()
        self.selectTarget = target
        self.preview.selecting = True

    def deselectPrevious(self):
        prevColumn = self.selectTarget.columnIndex
        if prevColumn >= 0:
            self.preview.unselect(prevColumn)

    def onSelect(self, column):
        self.deselectPrevious()
        self.selectTarget.select(column)
        self.checkButtonEnabled()

    def onDeselect(self, column):
        for frame in self.selectionFrames:
            if frame.columnIndex == column:
                frame.unselect()
                if frame is self.selectTarget:
                    self.checkButtonEnabled()
                    return
                break
        self.deselectPrevious()
        self.preview.select(column)
        self.selectTarget.select(column)
        self.propagateColumnColor(column, self.preview.columnFrames[column].cget('bg'))
        self.checkButtonEnabled()

    def propagateColumnColor(self, column, color):
        for frame in self.selectionFrames:
            if frame.columnIndex == column:
                frame.descriptionLabel.config(fg=color)
                break

class PreviewFrame(tk.Frame):
    def __init__(self, inputSpreadsheet: Spreadsheet, controller: ColumnSelectGUI, previewRowCount: int = 10):
        super().__init__(controller.window)
        self.controller = controller
        self.selecting = False
        self.selectedColumns = set()
        self.colorUsages = [0] * len(SELECTION_COLORS)

        previewRows = [inputSpreadsheet.getHeaders()]
        for r, row in itertools.islice(inputSpreadsheet.getRows(), previewRowCount - 1):
            previewRows.append(row)

        # count columns
        maxColumns = max(len(row) for row in previewRows)
        #print('maxColumns:', maxColumns)

        # make a frame for each column
        self.columnFrames = []
        for c in range(maxColumns):
            frame = tk.Frame(self)
            frame.grid(row=0, column=c)
            frame.bind('<Enter>', self.onColumnHoverEnter)
            frame.bind('<Leave>', self.onColumnHoverLeave)
            frame.bind('<Button-1>', self.onColumnClick)
            colLabel = tk.Label(frame, text=get_column_letter(c + 1), font='-weight bold')
            colLabel.grid(row=0, column=0)
            sep = tk.Frame(frame, highlightbackground='black', highlightthickness=1)
            sep.grid(row=1, column=0, sticky=tk.EW)
            self.columnFrames.append(frame)

        # fill in columns
        for r, row in enumerate(previewRows):
            # pad row with empty cells
            row = list(row) + [''] * (maxColumns - len(row))
            for c, cell in enumerate(row):
                #print('v', cell)
                cell = tk.Label(self.columnFrames[c], text=cell)
                cell.grid(row=r + 2, column=0, ipadx=10)
                cell.bind('<Button-1>', self.onColumnClick)

        # setup column colors
        for column in self.columnFrames:
            self.resetColumnBg(column)

    def pickColor(self):
        minIndex = 0
        minUsage = self.colorUsages[0]
        for i, usage in enumerate(self.colorUsages):
            if usage < minUsage:
                minIndex = i
                minUsage = usage
            if usage == 0:
                break
        self.colorUsages[minIndex] += 1
        return SELECTION_COLORS[minIndex]

    def returnColor(self, color):
        self.colorUsages[SELECTION_COLORS.index(color)] -= 1

    def setColumnFg(self, column, color):
        for child in column.winfo_children():
            if child.winfo_class() == 'Label':
                child.config(fg=color)

    def setColumnBg(self, column, color):
        column.config(bg=color)
        for child in column.winfo_children():
            if child.winfo_class() == 'Label':
                child.config(bg=color)

    def resetColumnBg(self, column):
        c = column.grid_info()['column']
        color = '#ffffff' if c % 2 == 0 else '#eeeeee'
        self.setColumnBg(column, color)

    def resetColumnFg(self, column):
        color = 'white' if column in self.selectedColumns else 'black'
        self.setColumnFg(column, color)

    def onColumnHoverEnter(self, event):
        if not self.selecting:
            return
        column = event.widget
        if column not in self.selectedColumns:
            self.setColumnFg(column, '#0066aa')
        else:
            self.setColumnFg(column, '#aaddff')

    def onColumnHoverLeave(self, event):
        column = event.widget
        self.resetColumnFg(column)

    def isSelected(self, c):
        column = self.columnFrames[c]
        return column in self.selectedColumns

    def select(self, column):
        if isinstance(column, int):
            column = self.columnFrames[column]
        if column in self.selectedColumns:
            return
        self.selectedColumns.add(column)
        self.setColumnBg(column, self.pickColor())
        self.setColumnFg(column, 'white')

    def unselect(self, column):
        if isinstance(column, int):
            column = self.columnFrames[column]
        self.selectedColumns.remove(column)
        self.returnColor(column.cget('bg'))
        self.resetColumnBg(column)
        self.resetColumnFg(column)

    def onColumnClick(self, event):
        if not self.selecting:
            return
        self.selecting = False

        column = event.widget
        if column.winfo_class() == 'Label':
            column = column.master

        c = column.grid_info()['column']

        if column in self.selectedColumns:
            print('unselected column', c)
            self.unselect(column)
            self.controller.onDeselect(c)
        else:
            print('selected column', c)
            self.controller.onSelect(c)
            self.select(column)
            self.controller.propagateColumnColor(c, column.cget('bg'))

class IndexSelectionFrame(tk.Frame):
    def __init__(self, indexInfo: IndexInfo, value: int, controller: ColumnSelectGUI):
        super().__init__(controller.window)
        self.indexInfo = indexInfo
        self.controller = controller
        self.columnIndex = value

        self.descriptionLabel = tk.Label(self, text=indexInfo.description + ': ')
        self.selectionBtn = ttk.Button(self, command=self.onSelectionBtnClick)
        self.descriptionLabel.pack(side=tk.LEFT)
        self.selectionBtn.pack(side=tk.LEFT)
        self.select(self.columnIndex)

    def select(self, index: int):
        if index < 0:
            self.unselect()
            return
        self.columnIndex = index
        self.selectionBtn.config(text=get_column_letter(index + 1))
        self.selectionBtn.state(['!disabled'])
        self.descriptionLabel.config(bg=self.controller.window.cget('bg'))

    def unselect(self):
        self.columnIndex = -1
        self.selectionBtn.config(text='(unselected)', style='')
        self.selectionBtn.state(['!disabled'])
        self.descriptionLabel.config(bg='#ffcccc', fg='black')

    def onSelectionBtnClick(self):
        self.controller.startSelect(self)
        self.selectionBtn.config(text='(selecting...)')
        self.selectionBtn.state(['disabled'])

    def cancelSelect(self):
        self.select(self.columnIndex)

if __name__ == '__main__':
    root = tk.Tk()
    root.title('tester')
    root.geometry('500x500')

    def onClick():
        indexGui = ColumnSelectGUI('donor-files/Originals/allDonors2022_CLEAN.xlsx')
        root.wait_window(indexGui.window)
        print(indexGui.result)
    tk.Button(root, text='column select test', command=onClick).pack()

    root.mainloop()